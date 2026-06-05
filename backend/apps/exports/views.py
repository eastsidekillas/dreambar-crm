import io
from datetime import datetime
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.db.models import Sum, F, Q
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.orders.models import Shift, Order, OrderItem, EntryTicket, UserProfile

# ── Style helpers ────────────────────────────────────────────────────────────

H_FILL    = PatternFill(start_color="1C3557", end_color="1C3557", fill_type="solid")
SH_FILL   = PatternFill(start_color="2D4E6D", fill_type="solid")
ACC_FILL  = PatternFill(start_color="B8922A", end_color="B8922A", fill_type="solid")
GREEN_F   = PatternFill(start_color="1A6632", fill_type="solid")
LIGHT_F   = PatternFill(start_color="EEF3F8", end_color="EEF3F8", fill_type="solid")
ALT_F     = PatternFill(start_color="F8F4EC", end_color="F8F4EC", fill_type="solid")

W_FONT  = Font(name='Calibri', color="FFFFFF", bold=True, size=11)
B_FONT  = Font(name='Calibri', bold=True, size=11)
N_FONT  = Font(name='Calibri', size=11)
S_FONT  = Font(name='Calibri', size=10, color="555555")

thin = Side(style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

RUB_FMT = '#,##0 ₽'
NUM_FMT = '#,##0'
DATE_FMT = 'DD.MM.YYYY'
TIME_FMT = 'DD.MM.YYYY HH:MM'


def hc(cell, text, fill=None, size=11):
    cell.value = text
    cell.font = Font(name='Calibri', color="FFFFFF", bold=True, size=size)
    cell.fill = fill or H_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER


def sc(cell, value, bold=False, fmt=None, align='left', fill=None):
    cell.value = value
    cell.font = B_FONT if bold else N_FONT
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill


def col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def emp_name(user):
    if user is None:
        return 'Неизвестно'
    profile = getattr(user, 'profile', None)
    return profile.get_display() if profile else (user.get_full_name() or user.username)


def emp_role(user):
    if user is None:
        return ''
    profile = getattr(user, 'profile', None)
    return profile.get_role_display() if profile else 'Официант'


# ── Sheet builders ───────────────────────────────────────────────────────────

def build_summary_sheet(wb, shifts):
    ws = wb.create_sheet(title='Сводная', index=0)
    ws.sheet_properties.tabColor = "B8922A"

    # Title
    ws.merge_cells('A1:I1')
    c = ws['A1']
    c.value = f'BAR DREAM — Сводный отчёт'
    c.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    c.fill = H_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    headers = ['Дата', 'День', 'Билеты (шт)', 'Билеты (₽)', 'Бар (₽)', 'Кухня (₽)', 'Кальян (₽)', 'ИТОГО (₽)', 'Заказов']
    widths   = [13,      12,    13,             14,            14,         14,           14,            16,           10]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hc(ws.cell(2, i), h)
        col_width(ws, i, w)
    ws.row_dimensions[2].height = 22

    totals = [0] * len(headers)
    for r, shift in enumerate(shifts.order_by('date'), 3):
        items = OrderItem.objects.filter(order__shift=shift, order__status='closed')
        t_n = shift.entry_tickets.count()
        t   = float(shift.entry_tickets.aggregate(s=Sum('price'))['s'] or 0)
        bar = float(items.filter(menu_item__category__type='bar').aggregate(s=Sum(F('unit_price')*F('quantity')))['s'] or 0)
        kitch = float(items.filter(menu_item__category__type='kitchen').aggregate(s=Sum(F('unit_price')*F('quantity')))['s'] or 0)
        hook = float(items.filter(menu_item__category__type='hookah').aggregate(s=Sum(F('unit_price')*F('quantity')))['s'] or 0)
        total = t + bar + kitch + hook
        orders = shift.orders.filter(status='closed').count()

        fill = ALT_F if r % 2 == 0 else None
        d = shift.date
        sc(ws.cell(r,1), d, fmt=DATE_FMT, fill=fill)
        sc(ws.cell(r,2), d.strftime('%A').capitalize(), fill=fill)
        sc(ws.cell(r,3), t_n, align='center', fill=fill)
        sc(ws.cell(r,4), t, fmt=RUB_FMT, align='right', fill=fill)
        sc(ws.cell(r,5), bar, fmt=RUB_FMT, align='right', fill=fill)
        sc(ws.cell(r,6), kitch, fmt=RUB_FMT, align='right', fill=fill)
        sc(ws.cell(r,7), hook, fmt=RUB_FMT, align='right', fill=fill)
        sc(ws.cell(r,8), total, bold=True, fmt=RUB_FMT, align='right', fill=fill)
        sc(ws.cell(r,9), orders, align='center', fill=fill)

        for j, v in enumerate([t_n, t, bar, kitch, hook, total, orders]):
            totals[j+2] += v

    tr = 3 + shifts.count()
    ws.cell(tr, 1).value = 'ИТОГО'
    ws.cell(tr, 1).font = W_FONT
    ws.cell(tr, 1).fill = ACC_FILL
    ws.cell(tr, 1).alignment = Alignment(horizontal='center')
    ws.merge_cells(f'A{tr}:B{tr}')
    for j, (v, fmt) in enumerate(zip(totals[2:], [NUM_FMT, RUB_FMT, RUB_FMT, RUB_FMT, RUB_FMT, RUB_FMT, NUM_FMT]), 3):
        c = ws.cell(tr, j)
        c.value = v
        c.font = W_FONT
        c.fill = ACC_FILL
        c.number_format = fmt
        c.alignment = Alignment(horizontal='right', vertical='center')
    ws.freeze_panes = 'A3'
    return ws


def build_employees_sheet(wb, shifts):
    ws = wb.create_sheet(title='Сотрудники')
    ws.sheet_properties.tabColor = "1C3557"

    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = 'BAR DREAM — Активность сотрудников'
    c.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    c.fill = H_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['Сотрудник', 'Должность', 'Заказов', 'Бар (₽)', 'Кухня (₽)', 'Кальян (₽)', 'Билетов (шт)', 'Билеты (₽)', 'ИТОГО (₽)']
    widths   = [22,           16,          10,         14,         14,           14,            13,             13,            16]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hc(ws.cell(2, i), h)
        col_width(ws, i, w)
    ws.row_dimensions[2].height = 22

    employees = User.objects.prefetch_related('profile').filter(
        Q(orders__shift__in=shifts) | Q(entryticket__shift__in=shifts)
    ).distinct()

    for r, emp in enumerate(employees, 3):
        emp_orders = Order.objects.filter(waiter=emp, shift__in=shifts, status='closed')
        emp_tickets = EntryTicket.objects.filter(created_by=emp, shift__in=shifts)
        items = OrderItem.objects.filter(order__in=emp_orders)
        bar   = float(items.filter(menu_item__category__type='bar').aggregate(t=Sum(F('unit_price')*F('quantity')))['t'] or 0)
        kitch = float(items.filter(menu_item__category__type='kitchen').aggregate(t=Sum(F('unit_price')*F('quantity')))['t'] or 0)
        hook  = float(items.filter(menu_item__category__type='hookah').aggregate(t=Sum(F('unit_price')*F('quantity')))['t'] or 0)
        t_n   = emp_tickets.count()
        t_rev = float(emp_tickets.aggregate(s=Sum('price'))['s'] or 0)
        total = bar + kitch + hook + t_rev

        fill = ALT_F if r % 2 == 0 else None
        sc(ws.cell(r,1), emp_name(emp), bold=True, fill=fill)
        sc(ws.cell(r,2), emp_role(emp), fill=fill)
        sc(ws.cell(r,3), emp_orders.count(), align='center', fill=fill)
        sc(ws.cell(r,4), bar, fmt=RUB_FMT, align='right', fill=fill)
        sc(ws.cell(r,5), kitch, fmt=RUB_FMT, align='right', fill=fill)
        sc(ws.cell(r,6), hook, fmt=RUB_FMT, align='right', fill=fill)
        sc(ws.cell(r,7), t_n, align='center', fill=fill)
        sc(ws.cell(r,8), t_rev, fmt=RUB_FMT, align='right', fill=fill)
        sc(ws.cell(r,9), total, bold=True, fmt=RUB_FMT, align='right', fill=fill)

    ws.freeze_panes = 'A3'
    return ws


def build_tickets_sheet(wb, shifts):
    ws = wb.create_sheet(title='Входные билеты')
    ws.sheet_properties.tabColor = "2ecc71"

    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = 'BAR DREAM — Входные билеты'
    c.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    c.fill = GREEN_F
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['Смена', '№ браслета', 'Цена', 'Дата/время', 'Сотрудник', 'Должность']
    widths   = [13,       15,           12,     18,            20,           14]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hc(ws.cell(2, i), h, fill=GREEN_F)
        col_width(ws, i, w)

    tickets = EntryTicket.objects.filter(shift__in=shifts).select_related(
        'shift', 'created_by', 'created_by__profile'
    ).order_by('shift__date', 'bracelet_number')

    for r, t in enumerate(tickets, 3):
        fill = ALT_F if r % 2 == 0 else None
        sc(ws.cell(r,1), t.shift.date, fmt=DATE_FMT, fill=fill)
        sc(ws.cell(r,2), t.bracelet_number, fill=fill)
        sc(ws.cell(r,3), float(t.price), fmt=RUB_FMT, align='right', fill=fill)
        sc(ws.cell(r,4), t.sold_at.replace(tzinfo=None), fmt=TIME_FMT, fill=fill)
        sc(ws.cell(r,5), emp_name(t.created_by), fill=fill)
        sc(ws.cell(r,6), emp_role(t.created_by), fill=fill)

    # Summary
    tr = 3 + tickets.count() + 1
    ws.cell(tr, 1).value = 'Итого билетов:'
    ws.cell(tr, 1).font = B_FONT
    sc(ws.cell(tr, 2), tickets.count(), bold=True, align='center')
    total_rev = tickets.aggregate(t=Sum('price'))['t'] or 0
    sc(ws.cell(tr, 3), float(total_rev), bold=True, fmt=RUB_FMT, align='right')

    ws.freeze_panes = 'A3'
    return ws


def build_orders_sheet(wb, shifts):
    ws = wb.create_sheet(title='Все заказы')
    ws.sheet_properties.tabColor = "7C3AED"

    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = 'BAR DREAM — Детальный список заказов'
    c.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    c.fill = PatternFill(start_color="4C1D95", fill_type="solid")
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['Смена', '№ заказа', 'Время', 'Стол/зона', 'Официант', 'Должность', 'Позиция', 'Объём', 'Кол-во', 'Сумма']
    widths   = [13,       10,         16,       12,           18,         14,           30,         10,     8,        12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hc(ws.cell(2, i), h, fill=PatternFill(start_color="4C1D95", fill_type="solid"))
        col_width(ws, i, w)

    orders = Order.objects.filter(
        shift__in=shifts, status='closed'
    ).select_related('waiter', 'waiter__profile', 'shift').prefetch_related(
        'items__menu_item'
    ).order_by('shift__date', 'created_at')

    row = 3
    for order in orders:
        for item in order.items.all():
            fill = ALT_F if row % 2 == 0 else None
            sc(ws.cell(row,1), order.shift.date, fmt=DATE_FMT, fill=fill)
            sc(ws.cell(row,2), order.id, align='center', fill=fill)
            sc(ws.cell(row,3), order.created_at.replace(tzinfo=None), fmt=TIME_FMT, fill=fill)
            sc(ws.cell(row,4), order.table_number or '—', fill=fill)
            sc(ws.cell(row,5), emp_name(order.waiter), fill=fill)
            sc(ws.cell(row,6), emp_role(order.waiter), fill=fill)
            sc(ws.cell(row,7), item.menu_item.name, fill=fill)
            sc(ws.cell(row,8), item.menu_item.volume or '—', fill=fill)
            sc(ws.cell(row,9), item.quantity, align='center', fill=fill)
            sc(ws.cell(row,10), float(item.subtotal), bold=True, fmt=RUB_FMT, align='right', fill=fill)
            row += 1

    ws.freeze_panes = 'A3'
    return ws


def build_shift_detail_sheet(wb, shift):
    # Unique, Excel-safe title (max 31 chars). Include id to avoid collisions.
    title = f"{shift.date.strftime('%d.%m.%Y')} №{shift.id}"[:31]
    ws = wb.create_sheet(title=title)

    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = f'Смена {shift.date.strftime("%d.%m.%Y")} ({shift.date.strftime("%A")})'
    c.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    c.fill = H_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Summary box
    items = OrderItem.objects.filter(order__shift=shift, order__status='closed')
    bar   = float(items.filter(menu_item__category__type='bar').aggregate(t=Sum(F('unit_price')*F('quantity')))['t'] or 0)
    kitch = float(items.filter(menu_item__category__type='kitchen').aggregate(t=Sum(F('unit_price')*F('quantity')))['t'] or 0)
    hook  = float(items.filter(menu_item__category__type='hookah').aggregate(t=Sum(F('unit_price')*F('quantity')))['t'] or 0)
    tick  = float(shift.entry_tickets.aggregate(t=Sum('price'))['t'] or 0)

    summary = [('Входные билеты', tick), ('Бар', bar), ('Кухня', kitch), ('Кальян', hook),
               ('ИТОГО', tick + bar + kitch + hook)]
    for i, (label, val) in enumerate(summary):
        r = 2 + i
        bold = label == 'ИТОГО'
        f = ACC_FILL if bold else None
        ws.cell(r, 1).value = label
        ws.cell(r, 1).font = W_FONT if bold else B_FONT
        if f: ws.cell(r, 1).fill = f
        ws.cell(r, 2).value = float(val)
        ws.cell(r, 2).number_format = RUB_FMT
        ws.cell(r, 2).font = W_FONT if bold else B_FONT
        if f: ws.cell(r, 2).fill = f

    col_width(ws, 1, 20)
    col_width(ws, 2, 16)

    # Tickets section
    row = 9
    ws.merge_cells(f'A{row}:H{row}')
    hc(ws.cell(row, 1), 'ВХОДНЫЕ БИЛЕТЫ', SH_FILL)
    ws.row_dimensions[row].height = 22
    row += 1
    for h, col in zip(['№', '№ браслета', 'Цена', 'Время', 'Сотрудник'], [1,2,3,4,5]):
        hc(ws.cell(row, col), h)
    row += 1
    for n, t in enumerate(shift.entry_tickets.select_related('created_by','created_by__profile').order_by('bracelet_number'), 1):
        sc(ws.cell(row,1), n, align='center')
        sc(ws.cell(row,2), t.bracelet_number)
        sc(ws.cell(row,3), float(t.price), fmt=RUB_FMT, align='right')
        sc(ws.cell(row,4), t.sold_at.replace(tzinfo=None), fmt=TIME_FMT)
        sc(ws.cell(row,5), emp_name(t.created_by))
        row += 1
    row += 1

    # Orders section
    ws.merge_cells(f'A{row}:H{row}')
    hc(ws.cell(row, 1), 'ЗАКАЗЫ (ДЕТАЛЬНО)', SH_FILL)
    ws.row_dimensions[row].height = 22
    row += 1
    for h, col in zip(['№ заказа', 'Время', 'Стол', 'Официант', 'Позиция', 'Объём', 'Кол-во', 'Сумма'], range(1,9)):
        hc(ws.cell(row, col), h)
    row += 1
    orders = Order.objects.filter(shift=shift, status='closed').select_related(
        'waiter','waiter__profile').prefetch_related('items__menu_item').order_by('created_at')
    for order in orders:
        for item in order.items.all():
            fill = ALT_F if row % 2 == 0 else None
            sc(ws.cell(row,1), order.id, align='center', fill=fill)
            sc(ws.cell(row,2), order.created_at.replace(tzinfo=None), fmt=TIME_FMT, fill=fill)
            sc(ws.cell(row,3), order.table_number or '—', fill=fill)
            sc(ws.cell(row,4), emp_name(order.waiter), fill=fill)
            sc(ws.cell(row,5), item.menu_item.name, fill=fill)
            sc(ws.cell(row,6), item.menu_item.volume or '—', fill=fill)
            sc(ws.cell(row,7), item.quantity, align='center', fill=fill)
            sc(ws.cell(row,8), float(item.subtotal), bold=True, fmt=RUB_FMT, align='right', fill=fill)
            row += 1

    for i, w in enumerate([8, 16, 10, 18, 28, 10, 8, 14], 1):
        col_width(ws, i, w)
    ws.freeze_panes = 'A3'
    return ws


# ── API Views ────────────────────────────────────────────────────────────────

class ExportShiftView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, shift_id):
        try:
            shift = Shift.objects.get(pk=shift_id)
        except Shift.DoesNotExist:
            return Response({'detail': 'Смена не найдена.'}, status=404)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        build_shift_detail_sheet(wb, shift)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"bardream_shift_{shift.date}.xlsx"
        resp = HttpResponse(buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp


class ExportReportView(APIView):
    """Full report: summary + employees + tickets + orders detail + per-shift sheets."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')

        shifts = Shift.objects.all()
        if date_from:
            shifts = shifts.filter(date__gte=date_from)
        if date_to:
            shifts = shifts.filter(date__lte=date_to)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        build_summary_sheet(wb, shifts)
        build_employees_sheet(wb, shifts)
        build_tickets_sheet(wb, shifts)
        build_orders_sheet(wb, shifts)
        for shift in shifts.order_by('-date'):
            build_shift_detail_sheet(wb, shift)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        now = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f"bardream_report_{now}.xlsx"
        resp = HttpResponse(buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
